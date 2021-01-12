import typing
import logging
import sys
import pandas as pd
from openpyxl import load_workbook
# from collections import namedtuple
#
# Columns = namedtuple('column', 'hostname, ip, status')

# Build the logger
log_file = "results.log"
file_handler = logging.FileHandler(log_file)
stream_handler = logging.StreamHandler(sys.stdout)
file_handler.setLevel(logging.DEBUG)
stream_handler.setLevel(logging.ERROR)
# formatting for loggers
file_handler_format = logging.Formatter('%(levelname)s - %(asctime)s - %(message)s')
stream_handler_format = logging.Formatter('%(levelname)s - %(asctime)s - %(message)s')
file_handler.setFormatter(file_handler_format)
stream_handler.setFormatter(stream_handler_format)

logger = logging.getLogger(__name__)
logger.addHandler(file_handler)
logger.addHandler(stream_handler)
logger.setLevel(logging.DEBUG)


class Column(typing.NamedTuple):
    hostname: str
    status: str
    ip: str


class ExcelProcessor:

    def __init__(self, spreadsheet, sheet, username, password, ignore_status=False):
        self.spreadsheet = spreadsheet
        self.username = username
        self.password = password
        self.ignore_status = ignore_status
        self.sheet = sheet
        # self.sheet = sheet_name
        # self.named_tuple: = Column()
        self.data = self.read_sheet()

    def read_sheet(self):
        # right method
        with pd.ExcelFile(self.spreadsheet) as xls:
            for sheet_name in xls.sheet_names:
                df = xls.parse(self.sheet)
                # df.reset_index()
                # df = pd.read_excel(xls, sheet_name=self.sheet)
                # print(df.head())
        return df

    def run_sheet_read(self):
        hosts = []
        for index, row in self.data.iterrows():
            data: Column = self.parse_case_sheet(Column, row)
            if not self.is_row_sane(data) or not self.process_row(data):
                continue
            # if not self.process_row(data):
            #     continue
            hosts.append(self.parse_hosts_for_netmiko(data.ip))
        return hosts

    def parse_hosts_for_netmiko(self, host, device_type='cisco_ios'):
        host = {
            'device_type': device_type,
            'host': host,
            'username': self.username,
            'password': self.password,
        }
        return host

    @staticmethod
    def is_row_sane(row):
        if row.hostname is None or row.ip is None:
            return False
        # if '.ngbutikk.net' not in row.hostname:
        #     return False
        return True

    def update_process_column(self, key, result):
        status = "success" if result else "failed"
        row = self.data.loc[self.data["ip"] == key, "status"] = status
        self.write_to_file()
        return row

    def update_sheet(self, key, value, column):
        row = self.data.loc[self.data["ip"] == key, column] = value
        self.write_to_file()
        return row

    def update_ports_column(self, key, port_info):
        ports = " ".join(port_info)
        self.update_sheet(key, ports, "ports")

    def write_to_file(self, index=False):
        # self.data.to_excel(self.spreadsheet, index=index)
        self.append_df_to_excel(index=index)

    def process_row(self, row):
        if self.ignore_status:
            return True
        if row.status == 'success':
            print(f"ignoring row {row}")
            logger.debug(f"ignoring row {row}")
            return False
        return True

    @staticmethod
    def clean_data(row):
        for key, value in row.items():
            if isinstance(value, str):
                value = value.strip()
            if pd.isna(value) or value == "" or value == "null":
                row[key] = None
        return row

    def parse_case_sheet(self, obj, row):
        return obj(**self.clean_data(row))

    def append_df_to_excel(self, startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        Parameters:
          startrow : upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
          truncate_sheet : truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
          to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                            [can be dictionary]

        Returns: None
        """
        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(self.spreadsheet, engine='openpyxl')

        try:
            # try to open an existing workbook
            writer.book = load_workbook(self.spreadsheet)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and self.sheet in writer.book.sheetnames:
                startrow = writer.book[self.sheet].max_row

            # truncate sheet
            if truncate_sheet and self.sheet in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(self.sheet)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(self.sheet, idx)

            # copy existing sheets
            writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass

        if startrow is None:
            startrow = 0

        # write out the new sheet
        # self.data.reset_index()
        self.data.to_excel(writer, self.sheet, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()
